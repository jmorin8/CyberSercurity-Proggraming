from pyhunter import PyHunter
from openpyxl import Workbook
import getpass


def Busqueda(organizacion):
    # Cantidad de resultados esperados de la búsqueda
    # El límite MENSUAL de Hunter es 50, cuidado!
    resultado = hunter.domain_search(
        company=organizacion, limit=1, emails_type='personal')
    return resultado


def data(datosEncontrados):
    for clave, valor in datosEncontrados.items():
        if isinstance(valor, list):
            dic = valor[0]
            data(dic)
        else:
            headers.append(clave)
            values.append(valor)
    return headers, values


def GuardarInformacion(headers, values, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")

    for i in range(1, 25):
        hoja.cell(i, 1, headers[i-1])
        hoja.cell(i, 2, values[i-1])

    libro.save("Hunter" + organizacion + ".xlsx")


headers = []
values = []

print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datosEncontrados = Busqueda(orga)

if datosEncontrados is None:
    exit()
else:
    data(datosEncontrados)
    GuardarInformacion(headers, values, orga)

